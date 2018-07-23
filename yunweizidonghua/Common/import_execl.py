#/usr/bin/env python
#-*- coding:utf-8 -*-
from openpyxl import load_workbook


def open_read_xls(path_file,filename_type):
    wb = load_workbook(filename=path_file)
    ws= wb.active
    max_rows = ws.max_row
    if filename_type == 'users':
        users = []
        for i in range(2,max_rows):

            col0 = ws.cell(row=i,column=1).value
            col1 = ws.cell(row=i,column=2).value
            col2 = ws.cell(row=i,column=3).value
            col3 = ws.cell(row=i,column=4).value
            col4 = ws.cell(row=i,column=5).value
            user_message = {'username':col0,'password':col1,'email':col2,'tel':col3,'user_type':col4}
            users.append(user_message)
        return users

    elif filename_type == 'hosts':
        hosts = []
        for i in range(2,max_rows):
            col0 = ws.cell(i,1).value
            col1 = ws.cell(i,2).value
            col2 = ws.cell(i,3).value
            col3 = ws.cell(i,4).value
            col4 = ws.cell(i,5).value
            host_message = {'hostname':col0,'ip':col1,'memory':col2,'disk':col3,'status_type':col4}
            hosts.append(host_message)

        return hosts


